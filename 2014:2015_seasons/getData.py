#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Mar 11 13:24:08 2019

@author: JayMessina
"""
import requests
import lxml.html as lh
import pandas as pd
import re
from bs4 import BeautifulSoup
import requests
import openpyxl
import datetime
import os

counting = 2
data = []
totalTakenOut = 0

flag = 0


def fillDataWithInjuries(directory, d, flag):
    book = openpyxl.Workbook(directory)
    counter = 0
    #currently have it run twice cuz some values miss
    for filename in os.listdir(directory):
        #filename = "Jordan Farmar_2014.xlsx"
        #print(filename)
        file = directory + "/" + filename
        n = filename.split("_")[0]
        if n not in d.keys():
            continue
        counter +=1
        r = ['GM', 'Date', 'Weight', 'MP', 'FGA', '3PA', 'FTA', 'ORB', 'DRB', 'AST', 'TO', 'Fouls', 'PTS', 'dist_feet', 'avg_speed', "post_ups", "drives", "Injury", "Type"]
        dataset = pd.read_excel(file, r=r)
        array = dataset.values
        #print(array)
        book = openpyxl.load_workbook(file)
        
        sheet = book['Sheet1']
        
        
        '''
        exp.
        Jordan Farmar
        dating = d["Jordan Farmar"] injury happened 2013-12-02
        s = sheet has date 2013-12-01
        d[n][s]
        
        '''
        
        
        sheet.cell(row=1, column=19).value = "Injury"
        sheet.cell(row=1, column=20).value = ""
        #print("here")
        #print(sheet.cell(row=1, column=17).value)
        
        for i in range(2, 90):
            #date value in sheet
            s = sheet.cell(row=i, column=3).value
            
            #value in dictionary is dating
            
            
            '''
            date from injury sheet, check date-1 day. if 0, subtract month and try 31, 30, or 28. 
            if true, then add it in that date -1.
            '''
            

            #if(s=="Date" or s ==None): 
            #   continue
            
            if s==None:
                continue
            
            
            nS = s.split("-")
            day = int(nS[2]) + 1
            nS[2] = str(day)
            if(len(nS[2])==1):
                nS[2] = "0" + nS[2]
            
            
            #calculate for overlap on days
            if(nS[1]=="01" or nS[1]=="03" or nS[1]=="05" or nS[1]=="07" or nS[1]=="08" or nS[1]=="10" or nS[1]=="12"):
                if(day==32):
                    nS[1] = str(int(nS[1])+1)
                    nS[2] = "01"
                if(nS[1]=="13"):
                    nS[1] = "01"
                    nS[0] = str(int(nS[0])+1)
                    
                if(len(nS[1])==1):
                        nS[1] = "0" + nS[1]
                    
                        
            elif(nS[1]=="04" or nS[1]=="06" or nS[1]=="09" or nS[1]=="11"):
                if(day==31):
                    nS[1] = str(int(nS[1])+1)
                    nS[2] = "01"
                    if(len(nS[1])==1):
                        nS[1] = "0" + nS[1]
                        
                
            elif(nS[1]=="02"):
                nS[1] = "03"
                nS[2] = "01"
                 
            newS = nS[0] + "-" + nS[1] + "-" + nS[2]
            
            '''
            print("newS ", newS)
            print("dating ", dating)
            print("s ", s)
            '''
            try:
                nextVal = d[n].get(newS)
                val = d[n].get(s)
                
                                    
                if(nextVal == "rest (DNP)" or nextVal == "placed on IL for rest"):
                    #print("here ", nextVal)
                    continue
                if(val==None):
                    val = d[n].get(newS)
                    
                    nextRow = sheet.cell(row=i+1, column=19).value
                    if(nextRow != newS and nextRow != None):
                        sheet.cell(row=i, column=19).value = val
                
                #put in after ran once
                if(flag ==1):
                    if(i>2):
                        minPlay = sheet.cell(row=i, column=5).value #==0
                        minPlayBefore = sheet.cell(row=i-1, column=5).value #!=0
            
                        if(minPlay==0 and minPlayBefore!=0):
                            sheet.cell(row=i-1, column=19).value = sheet.cell(row=i, column=19).value
                
   
                else:
                    sheet.cell(row=i, column=19).value = val

            except:
                continue
            
        book.save(file)
        #exit(0) 
         

def loadDict(file):
    names = ["Date", "Team", "Acquired", "Relinquished", "Notes"]
    dataset = pd.read_excel(file, names=names)
    array = dataset.values
    d = {}
    for x in array:
        date = str(x[0])
        name = x[3]
        injury = str(x[4])
        if(type(name)==float): 
            continue
        #split from first space
        name = name.split(' ', 1)[1]
        #split from •
        name = name.split(' ', 1)[1]
        name = name.replace(".", "")
        date = date.split(' ')[0]
        c = {name: {date:injury}}
        for k in d.keys():
             if name==k:
                 #print(d[name])
                 c[name].update(d[name])
                 break
                 #c = {name: {d[name]}, {{date:injury}}}
                # d[name].append({date:injury})
             
                
        d.update(c)
        
    return d
    
def loadInjured(beginDate, endDate, year):
    bad = [" illness (DNP)", " illness (DTD)", " rest (DNP)", " rest (DTD)", " returned to lineup", 
           " sports hernia (out indefinitely)", " surgery to repair sports hernia (out for season)", 
           " illness / upper respiratory infection (DTD)", " flu (DTD)", " flu (DNP)", " stomach flu (DNP)", 
           " flu / stomach virus (DTD)",  " stomach flu (DTD)", " flu/illness (DTD)", " stomach illness (DNP)", 
           " upper respiratory infection (DTD)", " upper respiratory infection / illness (DTD)", 
           " allergic reaction / illness (DTD)", " upper respiratory infection (DNP)", " skin infection (DNP)",
           " sinus infection (DNP)", " surgery on right elbow to remove staph infection", " flu / stomach virus (DNP)",
           " stomach flu (DTD)", " head coach returned to team", " vertigo (DNP)", " placed on IL for rest",
           " illness (out for season)", " upper respiratory illness (DTD)", " ingrown toenail (DNP)",
           " retired (effective at the end of the season)"]
    
    for i in range(0, 2000, 25):
        num = str(i)
        url = 'https://www.prosportstransactions.com/basketball/Search/SearchResults.php?Player=&Team=&BeginDate=' + beginDate + '&EndDate=' + endDate + '&InjuriesChkBx=yes&Submit=Search&start=' + num
        try:
            #Create a handle, page, to handle the contents of the website
            a = requests.get(url)
            soup = BeautifulSoup(a.text, 'lxml')
            #page = requests.get(url)
            
        except:
            break
      
        tr = soup.findAll('tr')
        
        bigL = []
        
        for x in tr:
            firstRow = 0
            items = x.findAll("td")
            #items = x.findAll("td")
        
            counter = 0
            L = []
            
            # filter by id with element, filter td that have the certain id
            for y in items:
                
                if((firstRow==0 and (y.string==None or y.string=="\xa0")) or (firstRow==0 and y.string=="\xa0Date")):
                    firstRow+=1
                    break
                if (y.string == "Previous"):
                    break
                L.append(y.string)
                counter+=1
                
                if (y.string in bad):
                    L = []
                    if (y.string != " returned to lineup" and y.string!=" head coach returned to team"):
                        global totalTakenOut
                        totalTakenOut+=1
                    
                
            #print(L)
            if(len(L)>0):
                bigL.append(L)
        '''
        global flag
        if (flag == 0):
            bigL.insert(0, ["Date", "Team", "Acquired", "Relinquished", "Notes"])
            flag+=1
        '''
        df = pd.DataFrame.from_records(bigL)
        df.dropna(thresh=3)
        data.append(df)
        '''
        if (i ==100):
            break
        '''
    result = pd.concat(data)
    file = "injured_data_" + year + ".xlsx"
    writer = pd.ExcelWriter(file)
    result.to_excel(writer,"Sheet1")
    writer.save()
    #print(totalTakenOut)
    
    #print(result)
    

def getSheet(link_first, link_second, version, name, dist_feet, avg_speed, posts, drives, year, avg_min):
    link = link_first + version + link_second
    a = requests.get(link)
    soup = BeautifulSoup(a.text, 'lxml')

    tb = soup.find("tbody")
    try:
        row = tb.findAll("tr")
        getW = soup.find('span', {'itemprop': 'weight'})
        weight = getW.text[0:3]
    except:
        v = int(version) + 1
        version = "0" + str(v)
        #print (link + " " + version)
        getSheet(link_first, link_second, version, name, dist_feet, avg_speed, posts, drives, year, avg_min)
        print(name)
        return
    global counting
    counting += 1
    print(name + " " + str(counting))
    labels = ['GM', 'Date', 'Weight', 'MP', 'FGA', '3PA', 'FTA', 'ORB', 'DRB', 'AST', 'TO', 'Fouls', 'PTS', 'dist_feet', 'avg_speed', "post_ups", "drives", "Injury"]
    
    bigL = []
    for x in row:
        items = x.findAll("td")
        counter = 0
        L = []
        for y in items:
            if(counter == 0):
                if(y.string=="None"):
                    print(y.string)
                    continue
                
            '''
            0 = game
            1 = date
            8 = minutes played
            10 = fga
            13 = 3fga
            16 = fta
            18 = offensive rebounds
            19 = def rebounds
            21 = assists
            24 = turnovers
            25 = personal fouls
            26 = points
            '''
            stats_to_ints = [0, 10, 13, 16, 18, 19, 21, 24, 25, 26]
            date = 1
            minutes = 8
            
            
            if (counter in stats_to_ints or counter == date or counter == minutes):
                if (counter in stats_to_ints):
                    x = y.string
                    if (x!=None):
                        L.append(int(x))
                    else:
                        L.append("")
                else:
                    if(counter==minutes):
                        s = y.text.split(":")
                        sInt0 = float(s[0])
                        sInt1 = float(s[1])
                        sInt1 /= 60
                        sInt1 = round(sInt1, 2)
                        
                        sInt0 += sInt1
                        L.append(sInt0)
                        
                        divisor = sInt0/avg_min
                        divisor = (round(divisor, 2))
                        
                    #date appending
                    else:
                        L.append(y.text)
                        L.append(int(weight))
            counter+=1
        if (len(L)>3):
            #float(str(round(answer, 2)))
            
            L.append(float(dist_feet)*divisor)
            L.append(float(avg_speed))
            L.append(float(posts)*divisor)
            L.append(float(drives)*divisor)
        if(len(L)==3):
            zero = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
            L.extend(zero)

        if(len(L)!=0):
            L.append("")
            #L.append("")
            bigL.append(L)
            

    df = pd.DataFrame.from_records(bigL, columns=labels)
    df.dropna()
    #print(df)
    filename = name + ".xlsx"
    if(year=="2014"):
        path = "bball_reference_2013-2014/" + filename
    else:
        path = "bball_reference_2014-2015/" + filename
        
    writer = pd.ExcelWriter(path)
    df.to_excel(writer,'Sheet1')
    writer.save()

def loadSportsVu(book, sheet, row_count, year):
    
    for i in range (2, row_count+1):
        s = sheet.cell(row=i, column=2).value
        if (s == None):
            continue
        #call function on s
        name = s[1:]
        
        name = re.sub(r'[^\w\s]','',name)
        
        if(name == "JJ Barea"):
            name = "Jose Barea"
        elif(name=="Henry Walker"):
            name = "Bill Walker"
        elif(name == "Luc Mbah a Moute"):
            name = "Luc Mbaha"
        elif(name == "Mo Williams"):
            name = "Maurice Williams"
        elif(name == "Nene"):
            name = "Nene Hilario"
            
        arr = name.split(" ")
        version = "01"
        #column 5 and 6
        avg_min = sheet.cell(row=i, column=4).value
        dist_feet = sheet.cell(row=i, column=5).value
        avg_speed = sheet.cell(row=i, column=6).value
        posts = sheet.cell(row=i, column=7).value
        drives = sheet.cell(row=i, column=8).value
        
        if (len(arr[1])>=5):
            link_first = ("https://www.basketball-reference.com/players/" + arr[0][0] + "/" + arr[1][0:5] + arr[0][0:2])
            #link_first += version
            link_second = "/gamelog/" + year + "/"
            link_first = link_first.lower()
            link_second = link_second.lower()
        else:
            link_first = ("https://www.basketball-reference.com/players/" + arr[0][0] + "/" + arr[1] + arr[0][0:2])
            link_second = "/gamelog/" + year + "/"
            link_first = link_first.lower()
            link_second = link_second.lower()
        f = name + "_" + year
        getSheet(link_first, link_second, version, f, dist_feet, avg_speed, posts, drives, year, avg_min)
        

def main():
    #steps
    #1. combined_fill.xlsx, which has every player who played more than avg 15 min. Scraped from SportsVU data. json to csv
    #2. get injured_edits from prosportstransaction into excel sheet, creating dictionary
    #3. loop through each spreadsheet and add the injured data to each player. run twice to check around it if reported on different day
    #4. run predictSeason.py to get actual values, separate program which takes in folder paths

    '''
    Step 1: script to get json to csv from sportsvu, get request from that website?
        - need to get name, avg min, avg speed, avg dist, post ups, drives to basket
    Step 2: have csv be xlsx file, and put them only if they avg more than 15 min
    Step 3: create a folder for the year (or cd to folder), for each name get their data from basketball reference and put in next couple of columns
        - data retried is date, min, FGA, FTA, points, assists
    Step 4: scrape pro sports transactions for all injury data and put in excel sheet. create dictionary. 
    Step 5: for each file in the folder, add injury data. only put in data if they played a few minutes (in game-injury). major injury if whole prediction window is filled with injury?
    Step 6: create training set, which will be all of season. then test set is each particular player. can predict probabilities for each game.
    
    '''
    '''
    a = requests.get("https://stats.nba.com/players/speed-distance/?sort=DIST_MILES&dir=1")
    soup = BeautifulSoup(a.text, 'lxml')
    url = 'https://stats.nba.com/stats/leaguedashptstats?College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick=&DraftYear=&GameScope=&Height=&LastNGames=0&LeagueID=00&Location=&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PerMode=PerGame&PlayerExperience=&PlayerOrTeam=Player&PlayerPosition=&PtMeasureType=PostTouch&Season=2014-15&SeasonSegment=&SeasonType=Regular+Season&StarterBench=&TeamID=0&VsConference=&VsDivision=&Weight='
    page = requests.get(url)
    
    tr_elements = doc.xpath('//table//thead//tr')
    rows = soup.findAll('table', {'class': 'row_summable sortable stats_table now_sortable'})

    tr = soup.find('div', {"class":"table_wrapper"}, "tbody")
    
    tb = soup.find("table")
    print(page)
    '''
    book1 = openpyxl.Workbook()
    book1 = openpyxl.load_workbook('2013-2014_combined_red.xlsx')
    sheet1 = book1.get_sheet_by_name('Sheet1')   
    row_count1 = sheet1.max_row
    year = "2014"
    #loadSportsVu(book1, sheet1, row_count1, year)
    
    
    book2 = openpyxl.Workbook()
    book2 = openpyxl.load_workbook('2014-2015_combined_red.xlsx')
    sheet2 = book2.get_sheet_by_name('Sheet1')   
    row_count2 = sheet2.max_row
    year = "2015"
    #loadSportsVu(book2, sheet2, row_count2, year)
    
    
    
    
    #for real time predictions, know beginning date then go to this
    #now = datetime.datetime.now()

    beginDate = "2013-10-25"
    endDate = "2014-04-18"
    year = "2014"
    #print (now.strftime("%Y-%m-%d"))
    
    #loadInjured(beginDate, endDate, year)
    file = "injured_data_2014.xlsx"
    dict2014 = {}
    dict2014 = loadDict(file)
    #print(dict2014)
    print(dict2014["Nate Robinson"])
    
    
    beginDate = "2014-10-25"
    endDate = "2015-04-18"
    year = "2015"
    
    #loadInjured(beginDate, endDate, year)
    file = "injured_data_2015.xlsx"
    dict2015 = {}
    dict2015 = loadDict(file)
    
    #print(dict2015["Paul Pierce"])
    
    ###########################
    #   load injury data      #
    ###########################
    
    
    directory = "bball_reference_2013-2014"
    #fillDataWithInjuries(directory, dict2014, 0)
    #fillDataWithInjuries(directory, dict2014, 1)
    
    
    directory = "bball_reference_2014-2015"
    #fillDataWithInjuries(directory, dict2015, 0)
    #fillDataWithInjuries(directory, dict2015, 1)

    
    
    
    

    

main()